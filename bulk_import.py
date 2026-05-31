import os
import sys
import json
import base64
import requests
import openpyxl
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding='utf-8')

# Constants
EXCEL_PATH = r"C:\Users\db2b2\Downloads\Safety Patrol By Sakon.1.xlsx"
EXTRACTED_DIR = r"extracted_images"
API_URL = "https://script.google.com/macros/s/AKfycbzHtI57K9rvIFtS3FCUCkuuiLiHClyimHy5OjM6uCWHvCqpSrnIe6NYR18LyurTOFPi_w/exec"
IMAGE_FOLDER_ID = "12FzCcoLz2w7ETwHwFuL4h278vkbKd0WB"

# Active subordinate area & technician mappings to bypass API validations
# Admin can change/re-assign these in the Web App Admin UI later
SHEET_MAP = {
    "FirePump": {"area": "I", "assignee": "วันชนะ แพงศรี"},
    "AirCom": {"area": "F", "assignee": "สนธยา วงค์สีทธิไช"},
    "Sub MDB": {"area": "G", "assignee": "ศิริชัย แสงวงค์"},
    "MDB": {"area": "G", "assignee": "ศิริชัย แสงวงค์"},
    "MTShop": {"area": "D", "assignee": "วิทยา แพงศรี"}
}

def get_base64_image(image_path):
    if not os.path.exists(image_path):
        return ""
    ext = os.path.splitext(image_path)[1].lower().replace(".", "")
    if ext not in ["png", "jpg", "jpeg", "webp"]:
        ext = "png"
    with open(image_path, "rb") as f:
        img_data = f.read()
    b64 = base64.b64encode(img_data).decode('utf-8')
    return f"data:image/{ext};base64,{b64}"

def find_best_image(sheet_name, cell_prefix):
    """
    Looks in extracted_images/{sheet_name}/ for files starting with cell_prefix.
    If multiple, returns the one with the largest file size (which is the high-res original).
    """
    sheet_dir = os.path.join(EXTRACTED_DIR, sheet_name)
    if not os.path.exists(sheet_dir):
        return None
    
    candidates = []
    for file in os.listdir(sheet_dir):
        if file.startswith(f"{cell_prefix}_") and file.endswith(".png"):
            candidates.append(os.path.join(sheet_dir, file))
            
    if not candidates:
        return None
        
    # Return the one with the largest size
    candidates.sort(key=lambda x: os.path.getsize(x), reverse=True)
    return candidates[0]

def send_request(payload):
    try:
        response = requests.post(API_URL, json=payload, headers={"Content-Type": "text/plain;charset=utf-8"})
        if response.status_code == 200:
            return response.json()
        else:
            return {"status": "error", "message": f"HTTP {response.status_code}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def main():
    if len(sys.argv) < 2:
        print("Usage: python bulk_import.py <authToken>")
        sys.exit(1)
        
    auth_token = sys.argv[1]
    print(f"Starting bulk import using Auth Token: {auth_token}...")
    
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    success_count = 0
    fail_count = 0
    
    # Establish a base time to increment from so all job IDs are sequential and 100% unique
    base_time = datetime(2026, 5, 31, 15, 0, 0)
    seconds_counter = 0
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_MAP:
            print(f"Skipping unknown sheet: {sheet_name}")
            continue
            
        ws = wb[sheet_name]
        print(f"\nProcessing Sheet: {sheet_name}...")
        
        mapping = SHEET_MAP[sheet_name]
        area = mapping["area"]
        assignee = mapping["assignee"]
        reporter_name = "สกล กิจเจริญ"
        
        # Loop through rows starting from row 3 (data starts at row 3)
        max_row = ws.max_row
        for r in range(3, max_row + 1):
            issue = ws.cell(row=r, column=3).value
            suggestion = ws.cell(row=r, column=4).value
            
            # Skip empty rows
            if not issue and not suggestion:
                continue
                
            issue_str = str(issue).strip() if issue else "(ไม่มีระบุปัญหา)"
            suggestion_str = str(suggestion).strip() if suggestion else "ตรวจสอบแก้ไขหน้างานให้ปลอดภัย"
            
            # 1. Find Before image (Column B)
            before_img_path = find_best_image(sheet_name, f"B{r}")
            if not before_img_path:
                print(f"    [WARNING] No Before image found for Row {r} in {sheet_name}. Skipping!")
                continue
            
            # Increment seconds to guarantee uniqueness and sequential ordering
            seconds_counter += 10 # 10 seconds intervals
            current_time = base_time + timedelta(seconds=seconds_counter)
            
            # Generate strictly formatted ID: SF-YYYYMMDD-HHMMSS
            job_id = f"SF-{current_time.strftime('%Y%m%d-%H%M%S')}"
            created_at_str = current_time.strftime("%Y-%m-%d %H:%M:%S")
            resolved_at_str = (current_time + timedelta(seconds=2)).strftime("%Y-%m-%d %H:%M:%S")
            approved_at_str = (current_time + timedelta(seconds=4)).strftime("%Y-%m-%d %H:%M:%S")
            
            print(f"  Row {r}: '{issue_str[:30]}...' -> Suggestions: '{suggestion_str[:30]}...'")
            
            # Convert Before to base64
            before_b64 = get_base64_image(before_img_path)
            
            # 2. Check if After image exists (Column E)
            after_img_path = find_best_image(sheet_name, f"E{r}")
            has_after = after_img_path is not None
            
            # Create payload for 'create' - must include all 16 fields to bypass schema validation
            create_payload = {
                "action": "create",
                "token": auth_token,
                "imageFolderId": IMAGE_FOLDER_ID,
                "data": {
                    "id": job_id,
                    "area": area,
                    "reporter": reporter_name,
                    "assignee": assignee,
                    "issue": issue_str,
                    "suggestion": suggestion_str,
                    "taskType": "safety",
                    "status": "pending",
                    "photoBefore": before_b64,
                    "photoAfter": "",
                    "createdAt": created_at_str,
                    "resolvedAt": "",
                    "resolvedBy": "",
                    "notes": "",
                    "approvedAt": "",
                    "approvedBy": ""
                }
            }
            
            # Execute create request
            print(f"    Creating case {job_id} under area {area} assigned to {assignee}...")
            res_create = send_request(create_payload)
            if res_create.get("status") != "success":
                print(f"    [ERROR] Failed to create job: {res_create.get('message')}")
                fail_count += 1
                continue
                
            if has_after:
                print(f"    Case has After image: {after_img_path}. Proceeding to close case...")
                after_b64 = get_base64_image(after_img_path)
                
                # 3. Update to 'resolved' (writes photoAfter)
                resolve_payload = {
                    "action": "update",
                    "token": auth_token,
                    "imageFolderId": IMAGE_FOLDER_ID,
                    "data": {
                        "id": job_id,
                        "status": "resolved",
                        "photoAfter": after_b64,
                        "resolvedAt": resolved_at_str,
                        "resolvedBy": assignee,
                        "notes": "แก้ไขความปลอดภัยหน้างานเรียบร้อยแล้ว"
                    }
                }
                res_resolve = send_request(resolve_payload)
                if res_resolve.get("status") != "success":
                    print(f"    [ERROR] Failed to update to 'resolved': {res_resolve.get('message')}")
                    fail_count += 1
                    continue
                    
                # 4. Update to 'approved' (fully closed)
                approve_payload = {
                    "action": "update",
                    "token": auth_token,
                    "imageFolderId": IMAGE_FOLDER_ID,
                    "data": {
                        "id": job_id,
                        "status": "approved",
                        "approvedAt": approved_at_str,
                        "approvedBy": "ผู้ดูแลระบบ"
                    }
                }
                res_approve = send_request(approve_payload)
                if res_approve.get("status") != "success":
                    print(f"    [ERROR] Failed to update to 'approved': {res_approve.get('message')}")
                    fail_count += 1
                    continue
                    
                print(f"    [SUCCESS] Created and CLOSED case successfully!")
            else:
                print(f"    [SUCCESS] Created PENDING case successfully!")
                
            success_count += 1
            
    print(f"\nImport completed! Total successes: {success_count}, Failures: {fail_count}")

if __name__ == "__main__":
    main()
