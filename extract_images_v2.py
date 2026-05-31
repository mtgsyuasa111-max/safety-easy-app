import openpyxl
import os
from PIL import Image as PILImage

excel_path = r"C:\Users\db2b2\Downloads\Safety Patrol By Sakon.1.xlsx"
wb = openpyxl.load_workbook(excel_path)

output_dir = "extracted_images"
os.makedirs(output_dir, exist_ok=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_dir = os.path.join(output_dir, sheet_name)
    os.makedirs(sheet_dir, exist_ok=True)
    
    images = ws._images
    print(f"Sheet: {sheet_name}, Images count: {len(images)}")
    
    for idx, img in enumerate(images):
        anchor = img.anchor
        if hasattr(anchor, '_from'):
            r = anchor._from.row + 1
            c = anchor._from.col + 1
            cell_name = f"{openpyxl.utils.get_column_letter(c)}{r}"
        else:
            cell_name = "Unknown"
            
        try:
            pil_img = PILImage.open(img.ref)
            filename = f"{cell_name}_img{idx}_{pil_img.width}x{pil_img.height}.png"
            file_path = os.path.join(sheet_dir, filename)
            pil_img.save(file_path)
        except Exception as e:
            print(f"  Error saving image {idx} in {sheet_name}: {e}")

print("Extraction completed!")
