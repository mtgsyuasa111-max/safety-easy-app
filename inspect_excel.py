import openpyxl
import pandas as pd
import os
import sys

# Set standard output encoding to utf-8
sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\db2b2\Downloads\Safety Patrol By Sakon.1.xlsx"

out_path = "excel_structure.txt"
with open(out_path, "w", encoding="utf-8") as out:
    out.write(f"Checking file size: {os.path.getsize(excel_path)} bytes\n")
    
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    out.write(f"Sheets in workbook: {wb.sheetnames}\n\n")
    
    for sheet_name in wb.sheetnames:
        out.write(f"=========================================\n")
        out.write(f"SHEET: {sheet_name}\n")
        out.write(f"=========================================\n")
        try:
            # Read first 15 rows
            df = pd.read_excel(excel_path, sheet_name=sheet_name, nrows=15)
            out.write("Columns:\n")
            out.write(str(df.columns.tolist()) + "\n\n")
            out.write("First 10 rows:\n")
            out.write(df.to_string() + "\n\n")
        except Exception as e:
            out.write(f"Error reading sheet {sheet_name}: {e}\n\n")

print("Done! Check excel_structure.txt")
