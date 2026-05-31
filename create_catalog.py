import openpyxl
import os

excel_path = r"C:\Users\db2b2\Downloads\Safety Patrol By Sakon.1.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

html_content = """<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Excel Safety Patrol Catalog</title>
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #0f172a; color: #e2e8f0; margin: 20px; }
        h1 { color: #10b981; }
        h2 { color: #f59e0b; border-bottom: 2px solid #1e293b; padding-bottom: 10px; margin-top: 40px; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; background: #1e293b; border-radius: 8px; overflow: hidden; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #334155; }
        th { background: #0f172a; color: #10b981; }
        img { max-width: 200px; max-height: 200px; border-radius: 4px; margin-right: 10px; border: 2px solid #334155; }
        .img-container { display: flex; flex-wrap: wrap; }
        .img-box { margin-right: 15px; text-align: center; }
        .img-label { font-size: 11px; color: #94a3b8; margin-top: 4px; }
    </style>
</head>
<body>
    <h1>Safety Patrol Excel Catalog</h1>
    <p>Visualizing all data and images from Safety Patrol By Sakon.1.xlsx</p>
"""

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    html_content += f"<h2>Sheet: {sheet_name}</h2>"
    html_content += "<table><thead><tr><th>Row</th><th>Images (Column B/E)</th><th>ปัญหา (Issue)</th><th>วิธีแก้ไข (Solution)</th></tr></thead><tbody>"
    
    # Extract images and associate with rows
    images = ws._images
    row_images = {}
    for idx, img in enumerate(images):
        anchor = img.anchor
        if hasattr(anchor, '_from'):
            r = anchor._from.row + 1
            c = anchor._from.col + 1
            col_letter = openpyxl.utils.get_column_letter(c)
            row_images.setdefault(r, []).append((idx, col_letter, img))
            
    # Read rows
    for r in range(3, ws.max_row + 1):
        issue = ws.cell(row=r, column=3).value
        solution = ws.cell(row=r, column=4).value
        
        if not issue and not solution and r not in row_images:
            continue
            
        html_content += f"<tr><td>{r}</td><td><div class='img-container'>"
        
        # Add images for this row
        imgs = row_images.get(r, [])
        # Sort images by index
        imgs.sort(key=lambda x: x[0])
        for img_idx, col_letter, img in imgs:
            # Find filename in extracted_images
            try:
                # Find size
                from PIL import Image as PILImage
                pil_img = PILImage.open(img.ref)
                w, h = pil_img.width, pil_img.height
            except Exception:
                w, h = 0, 0
                
            filename = f"extracted_images/{sheet_name}/{col_letter}{r}_img{img_idx}_{w}x{h}.png"
            # Relative path from where the html file will be run
            html_content += f"""
                <div class='img-box'>
                    <img src='{filename}' alt='Image {img_idx}'>
                    <div class='img-label'>Img {img_idx} ({col_letter}{r})<br>{w}x{h}</div>
                </div>
            """
        
        html_content += f"</div></td><td>{issue or ''}</td><td>{solution or ''}</td></tr>"
        
    html_content += "</tbody></table>"

html_content += """
</body>
</html>
"""

with open("catalog.html", "w", encoding="utf-8") as f:
    f.write(html_content)

print("Done! Open catalog.html")
