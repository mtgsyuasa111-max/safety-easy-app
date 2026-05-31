import openpyxl
from PIL import Image
import numpy as np
import os

excel_path = r"C:\Users\db2b2\Downloads\Safety Patrol By Sakon.1.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

def are_different(img_ref1, img_ref2):
    try:
        im1 = Image.open(img_ref1).convert('RGB')
        im2 = Image.open(img_ref2).convert('RGB')
        
        # Resize smaller to match larger, or vice versa
        if im1.size != im2.size:
            im2_resized = im2.resize(im1.size)
        else:
            im2_resized = im2
            
        arr1 = np.array(im1)
        arr2 = np.array(im2_resized)
        mse = np.mean((arr1 - arr2) ** 2)
        return mse > 1000 # If MSE > 1000, they are different images
    except Exception as e:
        print(f"Error comparing: {e}")
        return True

out_path = "image_pairs_analysis.txt"
with open(out_path, "w", encoding="utf-8") as out:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.write(f"\n=========================================\n")
        out.write(f"SHEET: {sheet_name}\n")
        out.write(f"=========================================\n")
        
        # Group images by row and column
        images = ws._images
        img_by_cell = {}
        for idx, img in enumerate(images):
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                r = anchor._from.row + 1
                c = anchor._from.col + 1
                cell_coord = f"{openpyxl.utils.get_column_letter(c)}{r}"
                img_by_cell.setdefault(cell_coord, []).append((idx, img))
                
        # For each cell, analyze images
        for cell_coord, imgs in sorted(img_by_cell.items(), key=lambda x: (int(x[0][1:]), x[0][0])):
            out.write(f"Cell {cell_coord}: {len(imgs)} image(s)\n")
            if len(imgs) == 1:
                out.write(f"  -> Single image (Img {imgs[0][0]}): size={imgs[0][1].width}x{imgs[0][1].height}\n")
            elif len(imgs) == 2:
                idx1, img1 = imgs[0]
                idx2, img2 = imgs[1]
                diff = are_different(img1.ref, img2.ref)
                if diff:
                    out.write(f"  -> TWO DIFFERENT IMAGES! (Img {idx1} vs Img {idx2})\n")
                    out.write(f"     Img {idx1}: {img1.width}x{img1.height}, Img {idx2}: {img2.width}x{img2.height}\n")
                else:
                    out.write(f"  -> Two copies of the SAME image (Img {idx1} is display, Img {idx2} is source)\n")
                    out.write(f"     Img {idx1}: {img1.width}x{img1.height}, Img {idx2}: {img2.width}x{img2.height}\n")
            elif len(imgs) > 2:
                out.write(f"  -> {len(imgs)} images in cell! Checking unique images...\n")
                # Compare all against first
                for i in range(len(imgs)):
                    for j in range(i+1, len(imgs)):
                        diff = are_different(imgs[i][1].ref, imgs[j][1].ref)
                        out.write(f"     Img {imgs[i][0]} vs Img {imgs[j][0]}: {'DIFFERENT' if diff else 'SAME'}\n")

print("Done! Check image_pairs_analysis.txt")
