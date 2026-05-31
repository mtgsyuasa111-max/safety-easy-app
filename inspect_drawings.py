import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\db2b2\Downloads\Safety Patrol By Sakon.1.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

out_path = "drawings_structure.txt"
with open(out_path, "w", encoding="utf-8") as out:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.write(f"Sheet: {sheet_name}\n")
        out.write(f"  Max row: {ws.max_row}, Max column: {ws.max_column}\n")
        
        # Print actual row 1 values (1-based index)
        row1_vals = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        row2_vals = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        out.write(f"  Row 1: {row1_vals}\n")
        out.write(f"  Row 2: {row2_vals}\n")
        
        images = ws._images
        out.write(f"  Number of images: {len(images)}\n")
        if len(images) > 0:
            for idx, img in enumerate(images):
                anchor = img.anchor
                out.write(f"    Image {idx}:\n")
                if isinstance(anchor, str):
                    out.write(f"      Anchor (str): {anchor}\n")
                else:
                    out.write(f"      Anchor type: {type(anchor).__name__}\n")
                    # Try to get from_
                    try:
                        if hasattr(anchor, 'from_'):
                            col = anchor.from_.col
                            row = anchor.from_.row
                            out.write(f"      From: Col={col}, Row={row} (0-based)\n")
                        if hasattr(anchor, 'to'):
                            col = anchor.to.col
                            row = anchor.to.row
                            out.write(f"      To: Col={col}, Row={row} (0-based)\n")
                    except Exception as e:
                        out.write(f"      Error getting details: {e}\n")
        out.write("-" * 50 + "\n")

print("Done! Check drawings_structure.txt")
