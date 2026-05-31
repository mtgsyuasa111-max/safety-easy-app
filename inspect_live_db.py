import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("live_database.xlsx")
print("Sheets in live database:", wb.sheetnames)

if "Users" in wb.sheetnames:
    ws = wb["Users"]
    print("\nUsers Tab contents:")
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"  Row {r}: {vals}")
else:
    # First sheet
    ws = wb.active
    print(f"\nActive Sheet ({ws.title}) contents:")
    for r in range(1, min(ws.max_row + 1, 15)):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"  Row {r}: {vals}")
