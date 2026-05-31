import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("live_database.xlsx")

for sheet_name in ["ชีต1", "ชีต2"]:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=========================================")
        print(f"SHEET: {sheet_name}")
        print(f"=========================================")
        for r in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if any(v is not None for v in row_vals):
                print(f"  Row {r}: {row_vals}")
