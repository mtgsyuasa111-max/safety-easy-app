import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\db2b2\Downloads\Safety Patrol By Sakon.1.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

out_path = "excel_detailed_inspect.txt"
with open(out_path, "w", encoding="utf-8") as out:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.write(f"\n=========================================\n")
        out.write(f"SHEET: {sheet_name}\n")
        out.write(f"=========================================\n")
        
        # Read all rows and columns to find text
        out.write("CELL VALUES:\n")
        for r in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            # If any cell in row is non-empty, print it
            if any(v is not None for v in row_vals):
                out.write(f"  Row {r}: {row_vals}\n")
        
        # Group images by row and column
        images = ws._images
        out.write(f"\nIMAGES IN SHEET ({len(images)}):\n")
        img_by_cell = {}
        for idx, img in enumerate(images):
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                r = anchor._from.row + 1 # 1-based row
                c = anchor._from.col + 1 # 1-based col
                cell_coord = f"{openpyxl.utils.get_column_letter(c)}{r}"
                img_by_cell.setdefault(cell_coord, []).append((idx, img))
            else:
                img_by_cell.setdefault("Unknown", []).append((idx, img))
        
        # Print grouped images
        for cell_coord, imgs_in_cell in sorted(img_by_cell.items(), key=lambda x: (x[0] == "Unknown", x[0])):
            out.write(f"  Cell {cell_coord}: {len(imgs_in_cell)} image(s)\n")
            for idx, img in imgs_in_cell:
                # Find size
                try:
                    width = img.width
                    height = img.height
                    out.write(f"    - Image {idx}: Size={width}x{height}\n")
                except Exception as e:
                    out.write(f"    - Image {idx}: Info error: {e}\n")

print("Done! Check excel_detailed_inspect.txt")
